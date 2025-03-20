import openpyxl

def invert_spreadsheet(input_file, output_file):
    """Reads an Excel file, transposes (swaps rows/columns), and saves a new file."""
    
    # Load the workbook and get the first sheet
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active
    
    # Get the max row and column count
    max_row = sheet.max_row
    max_col = sheet.max_column

    # Create a new workbook and sheet
    wb_new = openpyxl.Workbook()
    sheet_new = wb_new.active

    # Swap rows and columns
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            sheet_new.cell(row=col, column=row).value = sheet.cell(row=row, column=col).value

    # Save the new transposed file
    wb_new.save(output_file)
    print(f"✅ Transposed spreadsheet saved as {output_file}")

if __name__ == "__main__":
    input_file = input("Enter the input Excel file path: ").strip()
    output_file = input("Enter the output Excel file name: ").strip()

    try:
        invert_spreadsheet(input_file, output_file)
    except FileNotFoundError:
        print("❌ ERROR: File not found. Please check the file path.")
    except Exception as e:
        print(f"❌ ERROR: {e}")
