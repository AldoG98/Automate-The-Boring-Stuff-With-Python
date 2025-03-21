import openpyxl
import csv
import os

def excel_to_csv_all_sheets(excel_file, output_folder):
    """Converts each sheet in an Excel workbook to a separate CSV file."""
    
    # Load the Excel workbook
    wb = openpyxl.load_workbook(excel_file)

    # Create output folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)

    # Loop through each sheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        csv_filename = f"{sheet_name}.csv"
        csv_path = os.path.join(output_folder, csv_filename)

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            for row in sheet.iter_rows(values_only=True):
                writer.writerow(list(row))

        print(f"✅ Saved: {csv_filename}")

    print("\n🎉 All sheets converted successfully.")

if __name__ == "__main__":
    excel_path = input("Enter the Excel file path: ").strip()
    output_dir = input("Enter output folder name (or press Enter for 'csv_output'): ").strip()

    if output_dir == "":
        output_dir = "csv_output"

    try:
        excel_to_csv_all_sheets(excel_path, output_dir)
    except FileNotFoundError:
        print("❌ ERROR: Excel file not found.")
    except Exception as e:
        print(f"❌ ERROR: {e}")
